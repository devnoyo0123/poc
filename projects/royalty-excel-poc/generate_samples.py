#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font, Alignment

# 타입 A: 음반협회 (헤더 1행부터)
def create_type_a():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "음반협회 저작권료"

    # 헤더
    headers = ["저작권료", "저작자", "작품명", "사용일", "판매량"]
    ws.append(headers)

    # 데이터
    data = [
        ["100000", "홍길동", "Song1", "2024-01-01", "1000"],
        ["150000", "김철수", "Song2", "2024-01-02", "1500"],
        ["200000", "이영희", "Song3", "2024-01-03", "2000"],
        ["120000", "박민수", "Song4", "2024-01-04", "1200"],
        ["180000", "최수진", "Song5", "2024-01-05", "1800"]
    ]

    for row in data:
        ws.append(row)

    wb.save("samples/type_a_music.xlsx")
    print("타입 A 생성 완료: samples/type_a_music.xlsx")

# 타입 B: 방송협회 (헤더 3행부터)
def create_type_b():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "방송협회 정산"

    # 1행: 타이틀
    ws.cell(row=1, column=1, value="방송협회 저작권료 정산 보고서")
    ws.cell(row=1, column=1).font = Font(bold=True)

    # 2행: 기간
    ws.cell(row=2, column=1, value="정산 기간: 2024년 1월 ~ 3월")

    # 3행: 헤더
    headers = ["지급액", "권리자", "프로그램명", "방송일", "방송시간"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)

    # 4행부터: 데이터
    data = [
        ["50000", "KBS", "뉴스9", "2024-01-01", "60"],
        ["75000", "MBC", "무한도전", "2024-01-05", "90"],
        ["100000", "SBS", "런닝맨", "2024-01-10", "85"],
        ["60000", "tvN", "유 퀴즈 온 더 블럭", "2024-01-15", "75"],
        ["80000", "JTBC", "아는 형님", "2024-01-20", "80"]
    ]

    for row_idx, row_data in enumerate(data, 4):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save("samples/type_b_broadcast.xlsx")
    print("타입 B 생성 완료: samples/type_b_broadcast.xlsx")

# 타입 C: 영화협회 (헤더 2행부터)
def create_type_c():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "영화협회 상영료"

    # 1행: 빈 행 (나중에 요약 등을 넣을 수 있음)
    ws.cell(row=1, column=1, value=" ")

    # 2행: 헤더
    headers = ["상영료", "제작사", "영화제목", "개봉일", "관객수"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)

    # 3행부터: 데이터
    data = [
        ["5000000", "CJ엔터테인먼트", "범죄도시4", "2024-01-01", "1000000"],
        ["3000000", "NEW", "12월의 warmth", "2024-01-10", "600000"],
        ["4000000", "롯데엔터테인먼트", "노량: 죽음의 바다", "2024-01-15", "800000"],
        ["3500000", "이안스토리", "서울의 봄", "2024-01-20", "700000"],
        ["4500000", "필름머신", "3일의 휴일", "2024-01-25", "900000"]
    ]

    for row_idx, row_data in enumerate(data, 3):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save("samples/type_c_movie.xlsx")
    print("타입 C 생성 완료: samples/type_c_movie.xlsx")

# 타입 D: 공연협회 (헤더 1행부터, 컬럼 순서 다름)
def create_type_d():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "공연협회 공연료"

    # 헤더 (컬럼 순서가 다름!)
    headers = ["공연자", "티켓판매", "공연일", "공연료", "공연명"]
    ws.append(headers)

    # 헤더 스타일
    for col in range(1, 6):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)

    # 데이터 (컬럼 순서에 맞춰서)
    data = [
        ["BTS", "5000000", "2024-01-01", "100000000", "BTS WORLD TOUR"],
        ["BLACKPINK", "4000000", "2024-01-10", "80000000", "BLACKPINK CONCERT"],
        ["아이브", "3000000", "2024-01-15", "60000000", "IVE SHOW"],
        ["세븐틴", "4500000", "2024-01-20", "90000000", "SEVENTEEN TOUR"],
        ["뉴진스", "3500000", "2024-01-25", "70000000", "NEWJEANS CONCERT"]
    ]

    for row in data:
        ws.append(row)

    wb.save("samples/type_d_performance.xlsx")
    print("타입 D 생성 완료: samples/type_d_performance.xlsx")

if __name__ == "__main__":
    create_type_a()
    create_type_b()
    create_type_c()
    create_type_d()
    print("\n모든 샘플 엑셀 파일 생성 완료!")
